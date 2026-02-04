import requests, datetime, time
from openpyxl import load_workbook

URL = "https://query1.finance.yahoo.com/v8/finance/chart/{}"
TICKERS = ["SPY","QQQ","XLK","XLF","IWF","IWD","TLT","GLD"]

def fetch(t):
    r = requests.get(
        URL.format(t),
        params={"range": "1y", "interval": "1d"},
        headers={"User-Agent": "Mozilla/5.0"},
        timeout=20
    )
    r.raise_for_status()
    j = r.json()

    closes = j["chart"]["result"][0]["indicators"]["quote"][0]["close"]
    closes = [x for x in closes if x is not None]

    if len(closes) < 130:
        return None, None, None

    last = closes[-1]
    r3 = last / closes[-63] - 1
    r6 = last / closes[-126] - 1
    return last, r3, r6


def main(xlsx):
    wb = load_workbook(xlsx)
    ws = wb["Data"]

    spy_last, spy_r3, spy_r6 = fetch("SPY")
    if spy_last is None:
        raise RuntimeError("SPY data unavailable")

    for i, t in enumerate(TICKERS, start=2):
        data = fetch(t)
        if data[0] is None:
            continue

        last, r3, r6 = data
        ws.cell(i, 2, last)
        ws.cell(i, 3, r3)
        ws.cell(i, 4, r6)
        ws.cell(i, 5, r6 - spy_r6)
        ws.cell(i, 6, datetime.datetime.now())

        time.sleep(0.2)

    wb.save(xlsx)


if __name__ == "__main__":
    import sys
    main(sys.argv[1])
