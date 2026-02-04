import requests, datetime, time
from openpyxl import load_workbook

URL = "https://query1.finance.yahoo.com/v8/finance/chart/{}"
TICKERS = ["SPY","QQQ","XLK","XLF","IWF","IWD","TLT","GLD"]

def fetch(t):
    r = requests.get(URL.format(t), params={"range":"1y","interval":"1d"},
                     headers={"User-Agent":"Mozilla/5.0"})
    j = r.json()
    c = [x for x in j["chart"]["result"][0]["indicators"]["quote"][0]["close"] if x]
    last = c[-1]
    r3 = last/c[-63]-1
    r6 = last/c[-126]-1
    return last,r3,r6

def main(xlsx):
    wb = load_workbook(xlsx)
    ws = wb["Data"]
    spy = fetch("SPY")
    for i,t in enumerate(TICKERS, start=2):
        last,r3,r6 = fetch(t)
        ws.cell(i,2,last)
        ws.cell(i,3,r3)
        ws.cell(i,4,r6)
        ws.cell(i,5,r6-spy[2])
        ws.cell(i,6,datetime.datetime.now())
        time.sleep(0.2)
    wb.save(xlsx)

if __name__ == "__main__":
    import sys
    main(sys.argv[1])
